"""The flow the UI drives: upload last month, generate this month, download it."""
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import main
from app.service import RosterService


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "service", RosterService(history_dir=tmp_path / "history",
                                                       model_dir=tmp_path / "models",
                                                       export_dir=tmp_path / "exports"))
    return TestClient(main.app)


def _upload(client, payload, name="june.xlsx"):
    return client.post("/api/history/upload", files={"file": (name, payload)})


def test_health_and_rules(client):
    assert client.get("/api/health").json()["status"] == "ok"
    rules = client.get("/api/rules").json()
    assert rules["off_days_per_week"]["Night"] == 3
    assert rules["min_per_client_shift"] == 2


def test_upload_reads_the_sheet_and_names_the_month_to_build(client, wide_workbook_bytes, team):
    body = _upload(client, wide_workbook_bytes).json()

    assert body["month"] == "2025-06"
    assert body["target_month"] == "2025-07"
    assert len(body["employees"]) == len(team)
    assert body["ready"] is True
    assert body["blockers"] == []
    assert body["training"]["n_history_months"] == 1


def test_generate_returns_a_valid_month_and_exports_it(client, wide_workbook_bytes):
    _upload(client, wide_workbook_bytes)
    body = client.post("/api/roster/generate", json={"time_limit_seconds": 2}).json()

    assert body["month"] == "2025-07"
    assert body["validation"]["ok"], body["validation"]["errors"][:3]
    assert body["header"][:2] == ["Name", "Client"]
    assert body["header"][2] == "Tue-01-Jul"
    assert len(body["rows"]) == 23
    assert all(row["reason"] for row in body["rows"])
    assert min(entry["min"] for entry in body["coverage"]) >= 1

    roster_id = body["meta"]["id"]
    excel = client.get(f"/api/roster/{roster_id}/export.xlsx")
    assert excel.status_code == 200
    assert "attachment" in excel.headers["content-disposition"]

    workbook = load_workbook(BytesIO(excel.content))
    assert workbook.sheetnames == ["Roster Jul-2025", "Coverage check", "Summary", "Model"]
    sheet = workbook["Roster Jul-2025"]
    assert [c.value for c in sheet[1]][:3] == ["Name", "Client", "Tue-01-Jul"]
    assert sheet.max_row == 24

    csv = client.get(f"/api/roster/{roster_id}/export.csv")
    assert csv.status_code == 200
    assert csv.text.splitlines()[0].startswith("Name,Client,Tue-01-Jul")


def test_generating_without_history_explains_what_is_missing(client):
    response = client.post("/api/roster/generate", json={})
    assert response.status_code == 400
    assert "uploaded" in response.json()["detail"]


def test_an_impossible_team_returns_the_reasons(client, team):
    import json

    thin = [{"employee": row["employee"], "last_month_shift": row["last_month_shift"],
             "client": ["Client A", "Client B"]} for row in team[:6]]
    _upload(client, json.dumps(thin).encode(), name="team.json")

    response = client.post("/api/roster/generate", json={"month": "2025-07"})
    assert response.status_code == 422
    reasons = response.json()["detail"]["reasons"]
    assert any("at least 8" in reason for reason in reasons)


def test_history_accumulates_and_can_be_pruned(client, wide_workbook_bytes, team, tmp_path):
    import json

    _upload(client, wide_workbook_bytes)
    client.post("/api/history/upload", files={"file": ("may.json", json.dumps(team).encode())},
                data={"month": "2025-05"})

    months = client.get("/api/history").json()["months"]
    assert [m["month"] for m in months] == ["2025-05", "2025-06"]
    assert months[1]["with_off_pattern"] == len(team)

    assert client.delete("/api/history/2025-05").status_code == 200
    assert [m["month"] for m in client.get("/api/history").json()["months"]] == ["2025-06"]
    assert client.delete("/api/history/1999-01").status_code == 404


def test_the_model_reports_what_it_learned(client, wide_workbook_bytes, team):
    import json

    for month, offset in (("2025-03", 0), ("2025-04", 1), ("2025-05", 2)):
        rows = [{"employee": r["employee"], "client": r["client"],
                 "last_month_shift": ["Morning", "Afternoon", "Evening", "Night"][
                     (["Morning", "Afternoon", "Evening", "Night"].index(r["last_month_shift"])
                      + offset) % 4]} for r in team]
        client.post("/api/history/upload",
                    files={"file": (f"{month}.json", json.dumps(rows).encode())},
                    data={"month": month})

    report = client.post("/api/train").json()
    assert report["n_history_months"] == 3
    assert report["shift_model"]["trained"] is True
    assert report["shift_model"]["top1_accuracy"] is not None
    assert 0 < report["shift_model"]["blend_weight"] <= 0.85


def test_the_monthly_cycle_closes_and_the_model_learns(client, wide_workbook_bytes):
    """Each month's export is next month's upload - the loop has to close.

    It also has to *learn*: one month shows no shift change, so the shift model
    only starts training once a second month arrives, and its weight rises from
    there.
    """
    _upload(client, wide_workbook_bytes)
    weights = []

    for expected_month in ("2025-07", "2025-08", "2025-09"):
        roster = client.post("/api/roster/generate",
                             json={"time_limit_seconds": 2}).json()
        assert roster["month"] == expected_month
        assert roster["validation"]["ok"], roster["validation"]["errors"][:3]
        weights.append(roster["meta"]["training"]["shift_model"]["blend_weight"])

        # Feed this month's Excel export straight back in, as the user would.
        export = client.get(f"/api/roster/{roster['meta']['id']}/export.xlsx")
        uploaded = _upload(client, export.content, name=f"{expected_month}.xlsx").json()
        assert uploaded["month"] == expected_month
        assert uploaded["ready"] is True
        assert uploaded["warnings"] == []
        assert len(uploaded["employees"]) == len(roster["rows"])
        assert uploaded["clients"] == roster["clients"]

    # First month: nothing to learn from. Then it trains, and its say grows.
    assert weights[0] == 0.0
    assert weights[1] > 0.0
    assert weights[2] > weights[1]
    assert client.get("/api/model").json()["shift_model"]["trained"] is True


def test_the_built_ui_is_found_in_either_dist_layout(tmp_path):
    """Angular 22 writes dist/<project>/browser; older builds wrote the folder
    itself. Serving must keep working across that change."""
    from app.main import find_ui_build

    dist = tmp_path / "button-app"
    (dist / "browser").mkdir(parents=True)

    assert find_ui_build(dist) is None                 # nothing built yet

    (dist / "index.html").write_text("legacy")
    assert find_ui_build(dist) == dist                 # pre-upgrade layout

    (dist / "browser" / "index.html").write_text("current")
    assert find_ui_build(dist) == dist / "browser"     # current layout wins
