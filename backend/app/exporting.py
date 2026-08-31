"""Excel export - the deliverable the scheduling team actually works from."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import OFF_LABEL, SHIFTS, WEEKDAYS
from .roster import GeneratedRoster

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SHIFT_FILLS = {
    "Morning": PatternFill("solid", fgColor="FFF2CC"),
    "Afternoon": PatternFill("solid", fgColor="FCE4D6"),
    "Evening": PatternFill("solid", fgColor="E2EFDA"),
    "Night": PatternFill("solid", fgColor="DDEBF7"),
    OFF_LABEL: PatternFill("solid", fgColor="E7E6E6"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_header(sheet, values: list[str]) -> None:
    sheet.append(values)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER


def _widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook(roster: GeneratedRoster) -> Workbook:
    workbook = Workbook()

    # ---- the requested table: Name | Client | Mon-01-Jul | ... ----------
    sheet = workbook.active
    sheet.title = f"Roster {roster.month.label.replace(' ', '-')}"[:31]
    _write_header(sheet, roster.header)
    for row in roster.rows:
        sheet.append([row.name, row.client_label] + row.cells)
        written = sheet[sheet.max_row]
        written[0].border = written[1].border = BORDER
        for cell in written[2:]:
            cell.fill = SHIFT_FILLS.get(cell.value, SHIFT_FILLS[OFF_LABEL])
            cell.alignment = CENTER
            cell.border = BORDER
    _widths(sheet, [22, 26] + [11] * len(roster.month.days))
    sheet.freeze_panes = "C2"

    # ---- proof that every client is covered every day -------------------
    cover = workbook.create_sheet("Coverage check")
    _write_header(cover, ["Client", "Shift", "Team size", "Min on any day"]
                  + [d.label for d in roster.month.days])
    for entry in roster.coverage:
        cover.append([entry.client, entry.shift, entry.headcount, entry.minimum]
                     + entry.per_day)
        written = cover[cover.max_row]
        written[1].fill = SHIFT_FILLS[entry.shift]
        for cell in written:
            cell.border = BORDER
        for cell in written[2:]:
            cell.alignment = CENTER
            if cell.value == 0:
                cell.fill = SHIFT_FILLS[OFF_LABEL]
    _widths(cover, [16, 12, 11, 15] + [11] * len(roster.month.days))
    cover.freeze_panes = "E2"

    # ---- one line per person, including why they got this shift ---------
    summary = workbook.create_sheet("Summary")
    _write_header(summary, ["Name", "Client", "Last month shift", "This month shift",
                            "Week off", "Off days / week", "Working days", "Off days",
                            "Shift score", "Week-off score", "Why"])
    for row in roster.rows:
        block = ", ".join(WEEKDAYS[(row.off_start + i) % 7] for i in range(row.off_length))
        summary.append([row.name, row.client_label, row.previous_shift or "-", row.shift,
                        block, row.off_length, row.working_days, row.off_days,
                        row.shift_score, row.off_score, row.reason])
        written = summary[summary.max_row]
        written[3].fill = SHIFT_FILLS[row.shift]
        for cell in written:
            cell.border = BORDER
    summary.append([])
    headcounts = "   |   ".join(
        f"{shift}: {sum(1 for r in roster.rows if r.shift == shift)}" for shift in SHIFTS)
    summary.append(["Headcount per shift", headcounts])
    summary["A" + str(summary.max_row)].font = Font(bold=True)
    _widths(summary, [22, 26, 17, 17, 18, 15, 13, 11, 12, 14, 60])
    summary.freeze_panes = "A2"

    # ---- what the model knew when it built this -------------------------
    model_sheet = workbook.create_sheet("Model")
    _write_header(model_sheet, ["Property", "Value"])
    for key, value in _model_rows(roster):
        model_sheet.append([key, value])
        for cell in model_sheet[model_sheet.max_row]:
            cell.border = BORDER
    _widths(model_sheet, [34, 70])

    return workbook


def _model_rows(roster: GeneratedRoster) -> list[tuple[str, str]]:
    meta = roster.meta or {}
    training = meta.get("training") or {}
    shift_model = training.get("shift_model") or {}
    off_model = training.get("off_model") or {}
    validation = roster.validation or {}

    def describe(model: dict) -> str:
        if not model.get("trained"):
            return model.get("evaluation", "not trained yet")
        return (f"{model.get('estimator')} - top-1 {model.get('top1_accuracy')}, "
                f"AUC {model.get('roc_auc')}, weight {model.get('blend_weight')} "
                f"({model.get('evaluation')})")

    rows = [
        ("Roster month", roster.month.label),
        ("Generated at", meta.get("generated_at", "")),
        ("History months learned from", ", ".join(training.get("months", [])) or "none"),
        ("Shift model", describe(shift_model)),
        ("Week-off model", describe(off_model)),
        ("Optimiser status", meta.get("solver_status", "")),
        ("Optimiser time (s)", meta.get("solve_seconds", "")),
        ("Preference score", meta.get("objective", "")),
        ("All rules validated", "yes" if validation.get("ok") else "NO - see errors"),
    ]
    for note in meta.get("notes", []):
        rows.append(("Note", note))
    for name, weight in (shift_model.get("top_features") or [])[:6]:
        rows.append(("Shift model feature", f"{name}: {weight:+.3f}"))
    for name, weight in (off_model.get("top_features") or [])[:6]:
        rows.append(("Week-off model feature", f"{name}: {weight:+.3f}"))
    return rows


def export_to_bytes(roster: GeneratedRoster) -> bytes:
    stream = BytesIO()
    build_workbook(roster).save(stream)
    return stream.getvalue()


def export_to_file(roster: GeneratedRoster, path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(roster).save(path)
    return path


def export_to_csv(roster: GeneratedRoster) -> str:
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(roster.header)
    for row in roster.rows:
        writer.writerow([row.name, row.client_label] + row.cells)
    return buffer.getvalue()
